#!/usr/bin/env python3
import argparse
import hashlib
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Dict, Iterable, List, Optional

import psycopg2
from openpyxl import load_workbook


TRANSLIT = {
    "а": "a",
    "б": "b",
    "в": "v",
    "г": "g",
    "д": "d",
    "е": "e",
    "ё": "e",
    "ж": "zh",
    "з": "z",
    "и": "i",
    "й": "y",
    "к": "k",
    "л": "l",
    "м": "m",
    "н": "n",
    "о": "o",
    "п": "p",
    "р": "r",
    "с": "s",
    "т": "t",
    "у": "u",
    "ф": "f",
    "х": "h",
    "ц": "ts",
    "ч": "ch",
    "ш": "sh",
    "щ": "sch",
    "ъ": "",
    "ы": "y",
    "ь": "",
    "э": "e",
    "ю": "yu",
    "я": "ya",
}

HEADER_MAP = {
    "Дата оплаты": "payment_date",
    "Статус оплаты": "payment_status",
    "Дата начисления": "accrual_date",
    "Статус начисления": "accrual_status",
    "Контрагент": "counterparty",
    "ИНН контрагента": "counterparty_inn",
    "Тип": "row_type",
    "Счет": "account_name",
    "№ Счета": "account_number",
    "Банк": "bank_name",
    "Бик": "bank_bic",
    "Юрлицо": "legal_entity",
    "ИНН юрлица": "legal_entity_inn",
    "Статья": "article",
    "Родительские статьи": "parent_articles",
    "Вид деятельности": "activity_type",
    "Назначение платежа": "payment_purpose",
    "Проекты": "project_name",
    "Сумма": "amount",
    "Валюта": "currency",
}

RAW_FIELDS = [
    "payment_date",
    "payment_status",
    "accrual_date",
    "accrual_status",
    "counterparty",
    "counterparty_inn",
    "account_name",
    "account_number",
    "bank_name",
    "bank_bic",
    "legal_entity",
    "legal_entity_inn",
    "article",
    "parent_articles",
    "activity_type",
    "payment_purpose",
    "project_name",
    "amount",
    "currency",
]

INHERITED_FIELDS = [
    "payment_date",
    "payment_status",
    "accrual_date",
    "accrual_status",
    "counterparty",
    "counterparty_inn",
    "account_name",
    "account_number",
    "bank_name",
    "bank_bic",
    "legal_entity",
    "legal_entity_inn",
    "article",
    "parent_articles",
    "activity_type",
    "payment_purpose",
    "project_name",
    "amount",
    "currency",
]


@dataclass
class PlanFactRow:
    source_file_name: str
    source_sheet_name: str
    source_row_number: int
    row_type: str
    raw_values: Dict[str, object]
    parent_source_row_number: Optional[int] = None
    has_split_children: bool = False

    @property
    def is_split_part(self) -> bool:
        return self.row_type == "Часть"

    @property
    def entry_role(self) -> str:
        if self.is_split_part:
            return "split_part"
        if self.has_split_children:
            return "split_parent"
        return "entry"

    @property
    def root_source_row_number(self) -> int:
        return self.parent_source_row_number or self.source_row_number


def clean_text(value: object) -> Optional[str]:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


def normalize_date(value: object) -> Optional[date]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def normalize_amount(value: object) -> Optional[Decimal]:
    if value is None or value == "":
        return None
    return Decimal(str(value))


def stringify_for_key(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value).strip()


def slugify(text: str) -> str:
    lowered = text.strip().lower()
    transliterated = "".join(TRANSLIT.get(char, char) for char in lowered)
    transliterated = transliterated.replace("%", "pct")
    transliterated = re.sub(r"[^a-z0-9]+", "_", transliterated)
    transliterated = re.sub(r"_+", "_", transliterated).strip("_")
    return transliterated or "business_unit"


def source_period_end(rows: Iterable[PlanFactRow]) -> Optional[date]:
    candidates: List[date] = []
    for row in rows:
        for field in ("payment_date", "accrual_date"):
            value = row.raw_values.get(field)
            if isinstance(value, date):
                candidates.append(value)
    return max(candidates) if candidates else None


def build_dedupe_key(payload: Dict[str, object]) -> str:
    fields = [
        "row_type",
        "entry_role",
        "payment_date",
        "payment_status",
        "accrual_date",
        "accrual_status",
        "counterparty",
        "counterparty_inn",
        "account_name",
        "account_number",
        "bank_name",
        "bank_bic",
        "legal_entity",
        "legal_entity_inn",
        "article",
        "parent_articles",
        "activity_type",
        "payment_purpose",
        "project_name",
        "amount",
        "currency",
        "raw_payment_date",
        "raw_payment_status",
        "raw_accrual_date",
        "raw_accrual_status",
        "raw_counterparty",
        "raw_counterparty_inn",
        "raw_account_name",
        "raw_account_number",
        "raw_bank_name",
        "raw_bank_bic",
        "raw_legal_entity",
        "raw_legal_entity_inn",
        "raw_article",
        "raw_parent_articles",
        "raw_activity_type",
        "raw_payment_purpose",
        "raw_project_name",
        "raw_amount",
        "raw_currency",
    ]
    raw = "||".join(stringify_for_key(payload.get(field)) for field in fields)
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


def read_workbook(path: Path) -> List[PlanFactRow]:
    workbook = load_workbook(path, data_only=False)
    worksheet = workbook[workbook.sheetnames[0]]
    header_row = 2
    headers = [worksheet.cell(header_row, col).value for col in range(1, 21)]

    rows: List[PlanFactRow] = []
    current_parent_row_number: Optional[int] = None

    for row_number in range(3, worksheet.max_row + 1):
        values = [worksheet.cell(row_number, col).value for col in range(1, 21)]
        if not any(value is not None for value in values):
            continue

        raw_values: Dict[str, object] = {}
        for header, value in zip(headers, values):
            field = HEADER_MAP.get(header)
            if field is None:
                continue
            if field in {"payment_date", "accrual_date"}:
                raw_values[field] = normalize_date(value)
            elif field == "amount":
                raw_values[field] = normalize_amount(value)
            else:
                raw_values[field] = clean_text(value)

        row_type = raw_values.get("row_type")
        if not row_type:
            continue

        planfact_row = PlanFactRow(
            source_file_name=path.name,
            source_sheet_name=worksheet.title,
            source_row_number=row_number,
            row_type=row_type,
            raw_values=raw_values,
        )

        if row_type == "Часть":
            planfact_row.parent_source_row_number = current_parent_row_number
        else:
            current_parent_row_number = row_number

        rows.append(planfact_row)

    children_by_parent = defaultdict(int)
    for row in rows:
        if row.parent_source_row_number is not None:
            children_by_parent[row.parent_source_row_number] += 1

    for row in rows:
        row.has_split_children = children_by_parent[row.source_row_number] > 0

    return rows


def ensure_schema(conn) -> None:
    with conn.cursor() as cur:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS dim_business_units (
                id BIGSERIAL PRIMARY KEY,
                business_unit_code TEXT NOT NULL UNIQUE,
                business_unit_name TEXT NOT NULL UNIQUE,
                source_system TEXT NOT NULL DEFAULT 'planfact',
                is_active BOOLEAN NOT NULL DEFAULT TRUE,
                created_at TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW()
            );
            """
        )
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS planfact_cashflow_entries (
                id BIGSERIAL PRIMARY KEY,
                source_system TEXT NOT NULL DEFAULT 'planfact',
                source_file_name TEXT NOT NULL,
                source_sheet_name TEXT NOT NULL,
                source_row_number INTEGER NOT NULL,
                root_source_row_number INTEGER NOT NULL,
                parent_source_row_number INTEGER,
                row_type TEXT NOT NULL,
                entry_role TEXT NOT NULL,
                has_split_children BOOLEAN NOT NULL DEFAULT FALSE,
                payment_date DATE,
                payment_status TEXT,
                accrual_date DATE,
                accrual_status TEXT,
                counterparty TEXT,
                counterparty_inn TEXT,
                account_name TEXT,
                account_number TEXT,
                bank_name TEXT,
                bank_bic TEXT,
                legal_entity TEXT,
                legal_entity_inn TEXT,
                article TEXT,
                parent_articles TEXT,
                activity_type TEXT,
                payment_purpose TEXT,
                project_name TEXT,
                business_unit_id BIGINT REFERENCES dim_business_units(id),
                amount NUMERIC,
                currency TEXT,
                raw_payment_date DATE,
                raw_payment_status TEXT,
                raw_accrual_date DATE,
                raw_accrual_status TEXT,
                raw_counterparty TEXT,
                raw_counterparty_inn TEXT,
                raw_account_name TEXT,
                raw_account_number TEXT,
                raw_bank_name TEXT,
                raw_bank_bic TEXT,
                raw_legal_entity TEXT,
                raw_legal_entity_inn TEXT,
                raw_article TEXT,
                raw_parent_articles TEXT,
                raw_activity_type TEXT,
                raw_payment_purpose TEXT,
                raw_project_name TEXT,
                raw_amount NUMERIC,
                raw_currency TEXT,
                source_period_end DATE,
                dedupe_key TEXT,
                loaded_at TIMESTAMP WITHOUT TIME ZONE NOT NULL DEFAULT NOW(),
                CONSTRAINT planfact_cashflow_entries_source_row_unique
                    UNIQUE (source_file_name, source_sheet_name, source_row_number)
            );
            """
        )
        cur.execute(
            """
            ALTER TABLE planfact_cashflow_entries
            ADD COLUMN IF NOT EXISTS source_period_end DATE;
            """
        )
        cur.execute(
            """
            ALTER TABLE planfact_cashflow_entries
            ADD COLUMN IF NOT EXISTS dedupe_key TEXT;
            """
        )
        cur.execute(
            """
            CREATE INDEX IF NOT EXISTS planfact_cashflow_entries_business_unit_idx
                ON planfact_cashflow_entries (business_unit_id);
            """
        )
        cur.execute(
            """
            CREATE INDEX IF NOT EXISTS planfact_cashflow_entries_payment_date_idx
                ON planfact_cashflow_entries (payment_date);
            """
        )
        cur.execute(
            """
            CREATE INDEX IF NOT EXISTS planfact_cashflow_entries_accrual_date_idx
                ON planfact_cashflow_entries (accrual_date);
            """
        )
        cur.execute(
            """
            CREATE INDEX IF NOT EXISTS planfact_cashflow_entries_root_source_row_idx
                ON planfact_cashflow_entries (root_source_row_number);
            """
        )
        cur.execute(
            """
            CREATE INDEX IF NOT EXISTS planfact_cashflow_entries_dedupe_key_idx
                ON planfact_cashflow_entries (source_system, dedupe_key);
            """
        )
        cur.execute(
            """
            WITH file_periods AS (
                SELECT
                    source_file_name,
                    MAX(COALESCE(accrual_date, payment_date)) AS max_report_date
                FROM planfact_cashflow_entries
                WHERE source_system = 'planfact'
                GROUP BY source_file_name
            )
            UPDATE planfact_cashflow_entries e
            SET source_period_end = fp.max_report_date
            FROM file_periods fp
            WHERE e.source_system = 'planfact'
              AND e.source_file_name = fp.source_file_name
              AND e.source_period_end IS DISTINCT FROM fp.max_report_date;
            """
        )
        cur.execute(
            """
            UPDATE planfact_cashflow_entries
            SET dedupe_key = md5(
                concat_ws('||',
                    COALESCE(row_type, ''),
                    COALESCE(entry_role, ''),
                    COALESCE(payment_date::text, ''),
                    COALESCE(payment_status, ''),
                    COALESCE(accrual_date::text, ''),
                    COALESCE(accrual_status, ''),
                    COALESCE(counterparty, ''),
                    COALESCE(counterparty_inn, ''),
                    COALESCE(account_name, ''),
                    COALESCE(account_number, ''),
                    COALESCE(bank_name, ''),
                    COALESCE(bank_bic, ''),
                    COALESCE(legal_entity, ''),
                    COALESCE(legal_entity_inn, ''),
                    COALESCE(article, ''),
                    COALESCE(parent_articles, ''),
                    COALESCE(activity_type, ''),
                    COALESCE(payment_purpose, ''),
                    COALESCE(project_name, ''),
                    COALESCE(amount::text, ''),
                    COALESCE(currency, ''),
                    COALESCE(raw_payment_date::text, ''),
                    COALESCE(raw_payment_status, ''),
                    COALESCE(raw_accrual_date::text, ''),
                    COALESCE(raw_accrual_status, ''),
                    COALESCE(raw_counterparty, ''),
                    COALESCE(raw_counterparty_inn, ''),
                    COALESCE(raw_account_name, ''),
                    COALESCE(raw_account_number, ''),
                    COALESCE(raw_bank_name, ''),
                    COALESCE(raw_bank_bic, ''),
                    COALESCE(raw_legal_entity, ''),
                    COALESCE(raw_legal_entity_inn, ''),
                    COALESCE(raw_article, ''),
                    COALESCE(raw_parent_articles, ''),
                    COALESCE(raw_activity_type, ''),
                    COALESCE(raw_payment_purpose, ''),
                    COALESCE(raw_project_name, ''),
                    COALESCE(raw_amount::text, ''),
                    COALESCE(raw_currency, '')
                )
            )
            WHERE source_system = 'planfact'
              AND dedupe_key IS NULL;
            """
        )
    conn.commit()


def upsert_business_units(conn, rows: Iterable[PlanFactRow]) -> Dict[str, int]:
    names = sorted(
        {
            row.raw_values.get("project_name")
            for row in rows
            if clean_text(row.raw_values.get("project_name"))
        }
    )
    with conn.cursor() as cur:
        for name in names:
            code = slugify(name)
            cur.execute(
                """
                INSERT INTO dim_business_units (business_unit_code, business_unit_name, source_system, is_active)
                VALUES (%s, %s, 'planfact', TRUE)
                ON CONFLICT (business_unit_code)
                DO UPDATE SET
                    business_unit_name = EXCLUDED.business_unit_name,
                    is_active = TRUE
                """,
                (code, name),
            )
        cur.execute(
            """
            SELECT id, business_unit_name
            FROM dim_business_units
            WHERE business_unit_name = ANY(%s)
            """,
            (names,),
        )
        mapping = {name: unit_id for unit_id, name in cur.fetchall()}
    conn.commit()
    return mapping


def effective_value(row: PlanFactRow, parent_row: Optional[PlanFactRow], field: str):
    value = row.raw_values.get(field)
    if value is not None:
        return value
    if parent_row is None:
        return None
    return parent_row.raw_values.get(field)


def upsert_entries(
    conn,
    rows: List[PlanFactRow],
    business_units: Dict[str, int],
    file_period_end: Optional[date],
) -> None:
    rows_by_number = {row.source_row_number: row for row in rows}
    payloads = []

    for row in rows:
        parent_row = rows_by_number.get(row.parent_source_row_number)
        effective = {field: effective_value(row, parent_row, field) for field in INHERITED_FIELDS}
        project_name = effective.get("project_name")
        business_unit_id = business_units.get(project_name) if project_name else None

        payload = {
                "source_file_name": row.source_file_name,
                "source_sheet_name": row.source_sheet_name,
                "source_row_number": row.source_row_number,
                "root_source_row_number": row.root_source_row_number,
                "parent_source_row_number": row.parent_source_row_number,
                "row_type": row.row_type,
                "entry_role": row.entry_role,
                "has_split_children": row.has_split_children,
                "payment_date": effective["payment_date"],
                "payment_status": effective["payment_status"],
                "accrual_date": effective["accrual_date"],
                "accrual_status": effective["accrual_status"],
                "counterparty": effective["counterparty"],
                "counterparty_inn": effective["counterparty_inn"],
                "account_name": effective["account_name"],
                "account_number": effective["account_number"],
                "bank_name": effective["bank_name"],
                "bank_bic": effective["bank_bic"],
                "legal_entity": effective["legal_entity"],
                "legal_entity_inn": effective["legal_entity_inn"],
                "article": effective["article"],
                "parent_articles": effective["parent_articles"],
                "activity_type": effective["activity_type"],
                "payment_purpose": effective["payment_purpose"],
                "project_name": project_name,
                "business_unit_id": business_unit_id,
                "amount": effective["amount"],
                "currency": effective["currency"],
                "raw_payment_date": row.raw_values.get("payment_date"),
                "raw_payment_status": row.raw_values.get("payment_status"),
                "raw_accrual_date": row.raw_values.get("accrual_date"),
                "raw_accrual_status": row.raw_values.get("accrual_status"),
                "raw_counterparty": row.raw_values.get("counterparty"),
                "raw_counterparty_inn": row.raw_values.get("counterparty_inn"),
                "raw_account_name": row.raw_values.get("account_name"),
                "raw_account_number": row.raw_values.get("account_number"),
                "raw_bank_name": row.raw_values.get("bank_name"),
                "raw_bank_bic": row.raw_values.get("bank_bic"),
                "raw_legal_entity": row.raw_values.get("legal_entity"),
                "raw_legal_entity_inn": row.raw_values.get("legal_entity_inn"),
                "raw_article": row.raw_values.get("article"),
                "raw_parent_articles": row.raw_values.get("parent_articles"),
                "raw_activity_type": row.raw_values.get("activity_type"),
                "raw_payment_purpose": row.raw_values.get("payment_purpose"),
                "raw_project_name": row.raw_values.get("project_name"),
                "raw_amount": row.raw_values.get("amount"),
                "raw_currency": row.raw_values.get("currency"),
                "source_period_end": file_period_end,
            }
        payload["dedupe_key"] = build_dedupe_key(payload)
        payloads.append(payload)

    with conn.cursor() as cur:
        cur.executemany(
            """
            INSERT INTO planfact_cashflow_entries (
                source_file_name,
                source_sheet_name,
                source_row_number,
                root_source_row_number,
                parent_source_row_number,
                row_type,
                entry_role,
                has_split_children,
                payment_date,
                payment_status,
                accrual_date,
                accrual_status,
                counterparty,
                counterparty_inn,
                account_name,
                account_number,
                bank_name,
                bank_bic,
                legal_entity,
                legal_entity_inn,
                article,
                parent_articles,
                activity_type,
                payment_purpose,
                project_name,
                business_unit_id,
                amount,
                currency,
                raw_payment_date,
                raw_payment_status,
                raw_accrual_date,
                raw_accrual_status,
                raw_counterparty,
                raw_counterparty_inn,
                raw_account_name,
                raw_account_number,
                raw_bank_name,
                raw_bank_bic,
                raw_legal_entity,
                raw_legal_entity_inn,
                raw_article,
                raw_parent_articles,
                raw_activity_type,
                raw_payment_purpose,
                raw_project_name,
                raw_amount,
                raw_currency,
                source_period_end,
                dedupe_key
            ) VALUES (
                %(source_file_name)s,
                %(source_sheet_name)s,
                %(source_row_number)s,
                %(root_source_row_number)s,
                %(parent_source_row_number)s,
                %(row_type)s,
                %(entry_role)s,
                %(has_split_children)s,
                %(payment_date)s,
                %(payment_status)s,
                %(accrual_date)s,
                %(accrual_status)s,
                %(counterparty)s,
                %(counterparty_inn)s,
                %(account_name)s,
                %(account_number)s,
                %(bank_name)s,
                %(bank_bic)s,
                %(legal_entity)s,
                %(legal_entity_inn)s,
                %(article)s,
                %(parent_articles)s,
                %(activity_type)s,
                %(payment_purpose)s,
                %(project_name)s,
                %(business_unit_id)s,
                %(amount)s,
                %(currency)s,
                %(raw_payment_date)s,
                %(raw_payment_status)s,
                %(raw_accrual_date)s,
                %(raw_accrual_status)s,
                %(raw_counterparty)s,
                %(raw_counterparty_inn)s,
                %(raw_account_name)s,
                %(raw_account_number)s,
                %(raw_bank_name)s,
                %(raw_bank_bic)s,
                %(raw_legal_entity)s,
                %(raw_legal_entity_inn)s,
                %(raw_article)s,
                %(raw_parent_articles)s,
                %(raw_activity_type)s,
                %(raw_payment_purpose)s,
                %(raw_project_name)s,
                %(raw_amount)s,
                %(raw_currency)s,
                %(source_period_end)s,
                %(dedupe_key)s
            )
            ON CONFLICT (source_file_name, source_sheet_name, source_row_number)
            DO UPDATE SET
                root_source_row_number = EXCLUDED.root_source_row_number,
                parent_source_row_number = EXCLUDED.parent_source_row_number,
                row_type = EXCLUDED.row_type,
                entry_role = EXCLUDED.entry_role,
                has_split_children = EXCLUDED.has_split_children,
                payment_date = EXCLUDED.payment_date,
                payment_status = EXCLUDED.payment_status,
                accrual_date = EXCLUDED.accrual_date,
                accrual_status = EXCLUDED.accrual_status,
                counterparty = EXCLUDED.counterparty,
                counterparty_inn = EXCLUDED.counterparty_inn,
                account_name = EXCLUDED.account_name,
                account_number = EXCLUDED.account_number,
                bank_name = EXCLUDED.bank_name,
                bank_bic = EXCLUDED.bank_bic,
                legal_entity = EXCLUDED.legal_entity,
                legal_entity_inn = EXCLUDED.legal_entity_inn,
                article = EXCLUDED.article,
                parent_articles = EXCLUDED.parent_articles,
                activity_type = EXCLUDED.activity_type,
                payment_purpose = EXCLUDED.payment_purpose,
                project_name = EXCLUDED.project_name,
                business_unit_id = EXCLUDED.business_unit_id,
                amount = EXCLUDED.amount,
                currency = EXCLUDED.currency,
                raw_payment_date = EXCLUDED.raw_payment_date,
                raw_payment_status = EXCLUDED.raw_payment_status,
                raw_accrual_date = EXCLUDED.raw_accrual_date,
                raw_accrual_status = EXCLUDED.raw_accrual_status,
                raw_counterparty = EXCLUDED.raw_counterparty,
                raw_counterparty_inn = EXCLUDED.raw_counterparty_inn,
                raw_account_name = EXCLUDED.raw_account_name,
                raw_account_number = EXCLUDED.raw_account_number,
                raw_bank_name = EXCLUDED.raw_bank_name,
                raw_bank_bic = EXCLUDED.raw_bank_bic,
                raw_legal_entity = EXCLUDED.raw_legal_entity,
                raw_legal_entity_inn = EXCLUDED.raw_legal_entity_inn,
                raw_article = EXCLUDED.raw_article,
                raw_parent_articles = EXCLUDED.raw_parent_articles,
                raw_activity_type = EXCLUDED.raw_activity_type,
                raw_payment_purpose = EXCLUDED.raw_payment_purpose,
                raw_project_name = EXCLUDED.raw_project_name,
                raw_amount = EXCLUDED.raw_amount,
                raw_currency = EXCLUDED.raw_currency,
                source_period_end = EXCLUDED.source_period_end,
                dedupe_key = EXCLUDED.dedupe_key,
                loaded_at = NOW()
            """,
            payloads,
        )
    conn.commit()


def cleanup_duplicates(conn) -> int:
    with conn.cursor() as cur:
        cur.execute(
            """
            WITH ranked AS (
                SELECT
                    id,
                    ROW_NUMBER() OVER (
                        PARTITION BY source_system, dedupe_key
                        ORDER BY source_period_end DESC NULLS LAST, loaded_at DESC, id DESC
                    ) AS rn
                FROM planfact_cashflow_entries
                WHERE source_system = 'planfact'
                  AND dedupe_key IS NOT NULL
            ),
            deleted AS (
                DELETE FROM planfact_cashflow_entries p
                USING ranked r
                WHERE p.id = r.id
                  AND r.rn > 1
                RETURNING p.id
            )
            SELECT COUNT(*) FROM deleted;
            """
        )
        deleted_rows = cur.fetchone()[0]
    conn.commit()
    return deleted_rows


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx-path", required=True)
    parser.add_argument("--database-url", required=True)
    args = parser.parse_args()

    path = Path(args.xlsx_path)
    rows = read_workbook(path)
    file_period_end = source_period_end(rows)

    conn = psycopg2.connect(args.database_url)
    try:
        ensure_schema(conn)
        business_units = upsert_business_units(conn, rows)
        upsert_entries(conn, rows, business_units, file_period_end)
        deleted_rows = cleanup_duplicates(conn)
    finally:
        conn.close()

    print(f"Imported {len(rows)} rows from {path.name}; removed {deleted_rows} duplicate rows")


if __name__ == "__main__":
    main()
