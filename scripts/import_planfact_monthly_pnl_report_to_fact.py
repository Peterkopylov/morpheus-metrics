#!/usr/bin/env python3
from __future__ import annotations

import argparse
import calendar
import csv
from datetime import date
from pathlib import Path

import psycopg2
from openpyxl import load_workbook
from psycopg2.extras import Json, execute_batch

from planfact_monthly_pnl_report_mapping import (
    DEFAULT_WORKBOOK_PATHS,
    FACT_ROW_MAPPING,
    MONTH_LABELS,
    PLANFACT_REFERENCE_DOC,
    WORKSHEET_NAME,
    business_unit_from_workbook_name,
    business_unit_from_project_label,
    channel_from_row_label,
    detect_layout,
    planfact_report_type_for_business_unit,
    should_alert_on_leaf_to_parent_change,
    should_import_leaf_only_planfact_row,
)


INSERT_SQL = """
INSERT INTO fact_metric_observation (
    metric_id,
    rule_id,
    source_system,
    source_record_key,
    source_run_id,
    business_unit,
    show_name,
    partner_name,
    channel_name,
    period_granularity,
    period_start,
    period_end,
    value_numeric,
    value_text,
    value_raw,
    currency_code,
    is_estimated,
    payload
)
VALUES (
    %(metric_id)s,
    NULL,
    'planfact',
    %(source_record_key)s,
    %(source_run_id)s,
    %(business_unit)s,
    NULL,
    NULL,
    %(channel_name)s,
    'month',
    %(period_start)s,
    %(period_end)s,
    %(value_numeric)s,
    NULL,
    %(value_raw)s,
    'RUB',
    FALSE,
    %(payload)s
)
ON CONFLICT (
    metric_id,
    source_system,
    business_unit,
    show_name_norm,
    partner_name_norm,
    channel_name_norm,
    period_granularity,
    period_start,
    period_end,
    source_record_key_norm
)
DO UPDATE SET
    value_numeric = EXCLUDED.value_numeric,
    value_raw = EXCLUDED.value_raw,
    payload = EXCLUDED.payload,
    loaded_at = NOW()
"""


DELETE_SQL = """
DELETE FROM fact_metric_observation
WHERE source_system = 'planfact'
  AND business_unit = %(business_unit)s
  AND payload ->> 'planfact_report_type' = %(planfact_report_type)s
"""


def month_end(period_start: date) -> date:
    return date(period_start.year, period_start.month, calendar.monthrange(period_start.year, period_start.month)[1])


def normalize_numeric(value) -> tuple[str, float]:
    if value in (None, "", "-"):
        return "-", 0.0
    if isinstance(value, (int, float)):
        return str(value), float(value)
    raw = str(value).replace("\xa0", "").replace(" ", "").replace(",", ".").strip()
    if raw in {"", "-"}:
        return str(value), 0.0
    return str(value), float(raw)


def pnl_node_name_for_path(label: str) -> str:
    stripped = label.strip()
    aliases = {
        "Переменные расходы": "Переменные",
        "Постоянные расходы": "Постоянные",
    }
    return aliases.get(stripped, stripped)


def parse_pnl_label(raw_label) -> tuple[int, str, str]:
    text = str(raw_label or "")
    leading_spaces = len(text) - len(text.lstrip(" "))
    depth = leading_spaces // 4
    stripped = text.strip()
    return depth, stripped, pnl_node_name_for_path(text)


def write_report(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "workbook",
                "business_unit",
                "status",
                "source_row_number",
                "source_label",
                "month_label",
                "metric_name",
                "value_raw",
                "value_numeric",
                "reason",
            ],
        )
        writer.writeheader()
        writer.writerows(rows)


def resolve_business_unit(workbook_path: Path, worksheet) -> str:
    explicit_business_unit = business_unit_from_workbook_name(workbook_path.name)
    if explicit_business_unit:
        return explicit_business_unit

    project_label = str(worksheet.cell(4, 2).value or "").strip()
    business_unit = business_unit_from_project_label(project_label)
    if business_unit:
        return business_unit

    stem = workbook_path.stem.lower()
    aliases = {
        "b2cmoscow": "b2c_moscow",
        "b2cspb": "b2c_spb",
        "b2b": "b2b",
        "franchise": "franchise",
        "general": "general",
    }
    for marker, candidate in aliases.items():
        if marker in stem:
            return candidate
    raise RuntimeError(f"Unable to resolve business unit for workbook {workbook_path}")


def collect_month_columns(worksheet) -> list[tuple[int, str, date]]:
    header_row, _ = detect_layout(worksheet)
    month_columns: list[tuple[int, str, date]] = []
    for col_idx in range(2, worksheet.max_column + 1):
        month_label = worksheet.cell(header_row, col_idx).value
        period_label = MONTH_LABELS.get(str(month_label))
        if not period_label:
            continue
        month_columns.append((col_idx, str(month_label), date.fromisoformat(period_label)))
    return month_columns


def default_workbooks() -> list[str]:
    return [str(path) for path in DEFAULT_WORKBOOK_PATHS]


def should_keep_period(period_start: date, month_start: date | None, month_end_value: date | None) -> bool:
    if month_start and period_start < month_start:
        return False
    if month_end_value and period_start > month_end_value:
        return False
    return True


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--database-url", required=True)
    parser.add_argument("--xlsx-path", action="append", dest="xlsx_paths")
    parser.add_argument(
        "--report-path",
        default="/Users/Peter/Documents/Morpheus Metrics/generated/planfact_monthly_pnl_report_to_fact_import_report.csv",
    )
    parser.add_argument("--month-start", help="Optional YYYY-MM-DD lower month bound")
    parser.add_argument("--month-end", help="Optional YYYY-MM-DD upper month bound")
    parser.add_argument("--delete-existing", action="store_true")
    args = parser.parse_args()

    workbook_paths = [Path(path) for path in (args.xlsx_paths or default_workbooks())]
    month_start = date.fromisoformat(args.month_start) if args.month_start else None
    month_end_value = date.fromisoformat(args.month_end) if args.month_end else None

    conn = psycopg2.connect(args.database_url)
    try:
        with conn:
            with conn.cursor() as cur:
                cur.execute("SELECT metric_id, metric_name FROM metric_catalogue")
                metric_ids = {metric_name: metric_id for metric_id, metric_name in cur.fetchall()}

                inserts = []
                report_rows: list[dict[str, str]] = []
                for workbook_path in workbook_paths:
                    source_run_id = workbook_path.name
                    wb = load_workbook(workbook_path, data_only=True)
                    ws = wb[WORKSHEET_NAME]
                    business_unit = resolve_business_unit(workbook_path, ws)
                    planfact_report_type = planfact_report_type_for_business_unit(business_unit)
                    month_columns = collect_month_columns(ws)
                    _, data_start_row = detect_layout(ws)
                    path_stack: list[str] = []
                    parsed_rows: list[dict[str, object]] = []

                    if args.delete_existing:
                        delete_sql = DELETE_SQL
                        delete_params = {
                            "business_unit": business_unit,
                            "planfact_report_type": planfact_report_type,
                        }
                        if month_start:
                            delete_sql += "\n  AND period_start >= %(month_start)s"
                            delete_params["month_start"] = month_start
                        if month_end_value:
                            delete_sql += "\n  AND period_start <= %(month_end)s"
                            delete_params["month_end"] = month_end_value
                        cur.execute(delete_sql, delete_params)

                    for row_number in range(data_start_row, ws.max_row + 1):
                        raw_label = ws.cell(row_number, 1).value
                        if not raw_label:
                            continue
                        depth, label, path_label = parse_pnl_label(raw_label)
                        path_stack = path_stack[:depth]
                        path_stack.append(path_label)
                        pnl_node_path = " > ".join(path_stack)
                        parsed_rows.append(
                            {
                                "row_number": row_number,
                                "raw_label": raw_label,
                                "depth": depth,
                                "label": label,
                                "pnl_node_path": pnl_node_path,
                            }
                        )

                    for idx, parsed_row in enumerate(parsed_rows):
                        row_number = int(parsed_row["row_number"])
                        raw_label = parsed_row["raw_label"]
                        depth = int(parsed_row["depth"])
                        label = str(parsed_row["label"])
                        pnl_node_path = str(parsed_row["pnl_node_path"])
                        next_depth = int(parsed_rows[idx + 1]["depth"]) if idx + 1 < len(parsed_rows) else None
                        has_children = next_depth is not None and next_depth > depth

                        if has_children and should_alert_on_leaf_to_parent_change(label):
                            raise RuntimeError(
                                "PlanFact P&L structure changed: "
                                f"row {row_number} label {label!r} in workbook {workbook_path.name} "
                                "now has children but was previously treated as a leaf row. "
                                "Update generated/pnl_structure_mapping_canonical.csv and "
                                "scripts/planfact_monthly_pnl_report_mapping.py before reimporting."
                            )

                        include_row, skip_reason = should_import_leaf_only_planfact_row(
                            label=label,
                            has_children=has_children,
                        )
                        if not include_row:
                            for _, month_label, _ in month_columns:
                                report_rows.append(
                                    {
                                        "workbook": workbook_path.name,
                                        "business_unit": business_unit,
                                        "status": "skipped",
                                        "source_row_number": str(row_number),
                                        "source_label": label,
                                        "month_label": month_label,
                                        "metric_name": "",
                                        "value_raw": "",
                                        "value_numeric": "",
                                        "reason": skip_reason,
                                    }
                                )
                            continue

                        metric_name = FACT_ROW_MAPPING.get(label)
                        if not metric_name:
                            for _, month_label, _ in month_columns:
                                report_rows.append(
                                    {
                                        "workbook": workbook_path.name,
                                        "business_unit": business_unit,
                                        "status": "skipped",
                                        "source_row_number": str(row_number),
                                        "source_label": label,
                                        "month_label": month_label,
                                        "metric_name": "",
                                        "value_raw": "",
                                        "value_numeric": "",
                                        "reason": "no_metric_mapping",
                                    }
                                )
                            continue

                        metric_id = metric_ids.get(metric_name)
                        if not metric_id:
                            raise RuntimeError(
                                f"Metric {metric_name!r} for label {label!r} is missing from metric_catalogue."
                            )
                        channel_name = channel_from_row_label(label) or None

                        for col_idx, month_label, period_start in month_columns:
                            if not should_keep_period(period_start, month_start, month_end_value):
                                continue
                            cell_value = ws.cell(row_number, col_idx).value
                            value_raw, value_numeric = normalize_numeric(cell_value)
                            record_key = (
                                f"planfact_monthly_pnl_report:{business_unit}:"
                                f"r{row_number}:{period_start.isoformat()}"
                            )
                            inserts.append(
                                {
                                    "metric_id": metric_id,
                                    "source_record_key": record_key,
                                    "source_run_id": source_run_id,
                                    "business_unit": business_unit,
                                    "show_name": None,
                                    "partner_name": None,
                                    "channel_name": channel_name,
                                    "period_start": period_start,
                                    "period_end": month_end(period_start),
                                    "value_numeric": value_numeric,
                                    "value_raw": value_raw,
                                    "payload": Json(
                                        {
                                            "planfact_report_type": planfact_report_type,
                                            "source_workbook": str(workbook_path),
                                            "source_sheet": WORKSHEET_NAME,
                                            "source_row_number": row_number,
                                            "source_label": label,
                                            "source_label_raw": str(raw_label),
                                            "pnl_node_path": pnl_node_path,
                                            "pnl_depth": depth + 1,
                                            "source_month_label": month_label,
                                            "business_unit": business_unit,
                                            "channel_name": channel_name,
                                            "reference_doc": PLANFACT_REFERENCE_DOC,
                                        }
                                    ),
                                }
                            )
                            report_rows.append(
                                {
                                    "workbook": workbook_path.name,
                                    "business_unit": business_unit,
                                    "status": "inserted",
                                    "source_row_number": str(row_number),
                                    "source_label": label,
                                    "month_label": month_label,
                                    "metric_name": metric_name,
                                    "value_raw": value_raw,
                                    "value_numeric": str(value_numeric),
                                    "reason": "",
                                }
                            )

                execute_batch(cur, INSERT_SQL, inserts, page_size=200)
                write_report(Path(args.report_path), report_rows)
                print(
                    f"inserted={len(inserts)} "
                    f"workbooks={len(workbook_paths)} "
                    f"report={args.report_path}"
                )
    finally:
        conn.close()


if __name__ == "__main__":
    main()
