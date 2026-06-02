#!/usr/bin/env python3
from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from planfact_monthly_pnl_report_mapping import (
    CALCULATED_ROWS,
    DEFAULT_WORKBOOK_PATHS,
    FACT_ROW_MAPPING,
    MONTH_LABELS,
    PLANFACT_HOW_COUNTED,
    PLANFACT_REFERENCE_DOC,
    REVENUE_DETAIL_ROWS,
    REMOVED_METRIC_NAMES,
    ROOT,
    WORKSHEET_NAME,
    business_unit_from_workbook_name,
    business_unit_from_project_label,
    detect_layout,
)


CATALOGUE_CSV = ROOT / "generated" / "metric_catalogue_canonical.csv"
SOURCES_CSV = ROOT / "generated" / "fact_metric_source_of_truth_canonical.csv"
REPORT_CSV = ROOT / "generated" / "planfact_monthly_pnl_fact_vs_calculated.csv"


def load_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def iter_workbooks() -> list[tuple[Path, str]]:
    resolved: list[tuple[Path, str]] = []
    for workbook_path in DEFAULT_WORKBOOK_PATHS:
        wb = load_workbook(workbook_path, data_only=True)
        ws = wb[WORKSHEET_NAME]
        business_unit = business_unit_from_workbook_name(workbook_path.name)
        if not business_unit:
            project_label = str(ws.cell(4, 2).value or "").strip()
            business_unit = business_unit_from_project_label(project_label)
        if not business_unit:
            raise RuntimeError(f"Unknown PlanFact workbook mapping for {workbook_path}")
        resolved.append((workbook_path, business_unit))
    return resolved


def rewrite_catalogue() -> tuple[int, int]:
    rows = load_csv(CATALOGUE_CSV)
    before = len(rows)
    updated = [row for row in rows if row["metric_name"] not in REMOVED_METRIC_NAMES]
    write_csv(CATALOGUE_CSV, updated, list(rows[0].keys()))
    return before, len(updated)


def build_source_rules_to_ensure(metric_key_by_name: dict[str, str]) -> list[dict[str, str]]:
    rules: dict[tuple[str, str], dict[str, str]] = {}
    for workbook_path, business_unit in iter_workbooks():
        wb = load_workbook(workbook_path, data_only=True)
        ws = wb[WORKSHEET_NAME]
        _, data_start_row = detect_layout(ws)
        for row_number in range(data_start_row, ws.max_row + 1):
            raw_label = ws.cell(row_number, 1).value
            if not raw_label:
                continue
            label = str(raw_label).strip()
            if label in CALCULATED_ROWS:
                continue
            if label in REVENUE_DETAIL_ROWS:
                continue
            metric_name = FACT_ROW_MAPPING.get(label)
            if not metric_name:
                continue
            metric_key = metric_key_by_name.get(metric_name)
            if not metric_key:
                raise RuntimeError(
                    f"Metric {metric_name!r} for label {label!r} is missing from canonical catalogue."
                )
            key = (metric_name, business_unit)
            if key in rules:
                continue
            rules[key] = {
                "metric_key": metric_key,
                "metric_name": metric_name,
                "business_unit": business_unit,
                "show_scope": "general",
                "partner_scope": "general",
                "channel_scope": "general",
                "frequency": "month",
                "source_role": "primary",
                "source_system": "planfact",
                "how_counted": PLANFACT_HOW_COUNTED,
                "reference_doc": PLANFACT_REFERENCE_DOC,
                "status_note": "",
                "source_row_ref": f"{workbook_path.name}:row_{row_number}",
            }
    return sorted(rules.values(), key=lambda row: (row["metric_name"], row["business_unit"]))


def rewrite_sources() -> tuple[int, int]:
    rows = load_csv(SOURCES_CSV)
    kept = []
    for row in rows:
        if row["source_system"] == "planfact" and row["reference_doc"] == PLANFACT_REFERENCE_DOC:
            continue
        if (
            row["source_system"] == "planfact"
            and row["metric_name"] == "Costs - Salary variable"
            and not row["how_counted"].strip()
        ):
            continue
        kept.append(row)
    metric_key_by_name = {row["metric_name"]: row["metric_key"] for row in load_csv(CATALOGUE_CSV)}
    added_rows = build_source_rules_to_ensure(metric_key_by_name)
    updated = kept + added_rows
    updated.sort(
        key=lambda row: (
            row["metric_name"],
            row["business_unit"],
            row["show_scope"],
            row["partner_scope"],
            row["channel_scope"],
            row["source_system"],
            row["frequency"],
        )
    )
    write_csv(SOURCES_CSV, updated, list(rows[0].keys()))
    return len(rows), len(updated)


def build_fact_vs_calculated_report() -> tuple[int, int, int]:
    report_rows: list[dict[str, str]] = []
    fact_count = 0
    calculated_count = 0
    unmapped_count = 0

    for workbook_path, business_unit in iter_workbooks():
        wb = load_workbook(workbook_path, data_only=True)
        ws = wb[WORKSHEET_NAME]
        header_row, data_start_row = detect_layout(ws)
        month_labels = [ws.cell(header_row, col).value for col in range(2, ws.max_column + 1) if ws.cell(header_row, col).value in MONTH_LABELS]
        for row_number in range(data_start_row, ws.max_row + 1):
            raw_label = ws.cell(row_number, 1).value
            if not raw_label:
                continue
            label = str(raw_label).strip()
            if label in FACT_ROW_MAPPING:
                classification = "fact"
                target = FACT_ROW_MAPPING[label]
                note = "Observed monthly P&L row; import into fact_metric_observation."
                fact_count += 1
            elif label in REVENUE_DETAIL_ROWS:
                classification = "excluded"
                target = ""
                note = "Revenue child row is skipped; import only the top subtotal row `Выручка`."
            elif label in CALCULATED_ROWS:
                classification = "calculated"
                target = ""
                note = "Derived percentage / margin row; do not import into raw fact layer."
                calculated_count += 1
            else:
                classification = "unmapped"
                target = ""
                note = "No mapping configured yet."
                unmapped_count += 1

            non_empty_months = 0
            for col_idx in range(2, 2 + len(month_labels)):
                value = ws.cell(row_number, col_idx).value
                if value not in (None, ""):
                    non_empty_months += 1

            report_rows.append(
                {
                    "workbook": workbook_path.name,
                    "business_unit": business_unit,
                    "source_row_number": str(row_number),
                    "source_label": label,
                    "classification": classification,
                    "target_metric_name": target,
                    "months_present": str(non_empty_months),
                    "note": note,
                }
            )

    write_csv(
        REPORT_CSV,
        report_rows,
        [
            "workbook",
            "business_unit",
            "source_row_number",
            "source_label",
            "classification",
            "target_metric_name",
            "months_present",
            "note",
        ],
    )
    return fact_count, calculated_count, unmapped_count


def main() -> None:
    catalogue_before, catalogue_after = rewrite_catalogue()
    sources_before, sources_after = rewrite_sources()
    fact_count, calculated_count, unmapped_count = build_fact_vs_calculated_report()
    print(
        f"catalogue_before={catalogue_before} "
        f"catalogue_after={catalogue_after} "
        f"sources_before={sources_before} "
        f"sources_after={sources_after} "
        f"fact_rows={fact_count} "
        f"calculated_rows={calculated_count} "
        f"unmapped_rows={unmapped_count} "
        f"report={REPORT_CSV}"
    )


if __name__ == "__main__":
    main()
